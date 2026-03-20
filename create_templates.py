"""Creates sample Excel templates for Purchase Register and Sales Register."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

DARK  = "1F3864"
MED   = "2E75B6"
WHITE = "FFFFFF"
THIN  = Side(style="thin")
BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr(ws, row, col, val, bg=DARK):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BDR

def sample(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal="left")
    c.border = BDR

def make_purchase_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Register"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "PURCHASE REGISTER – Sample Template for GST Reconciliation Tool"
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = ("Delete this row and Row 1 before loading. Keep headers in Row 1 only.  "
               "You may add more rows. Date format: DD-MM-YYYY")
    c.fill = PatternFill("solid", fgColor=MED)
    c.font = Font(bold=True, color=WHITE, size=9)
    c.alignment = Alignment(horizontal="center")

    headers = ["GSTIN", "Supplier Name", "Invoice No", "Invoice Date",
               "Invoice Value", "Taxable Value", "IGST", "CGST", "SGST", "CESS",
               "Remarks"]
    for col, h in enumerate(headers, 1):
        hdr(ws, 3, col, h)

    samples = [
        ["27AABCU9603R1ZN", "ABC Pvt Ltd",        "INV/2024/001", "05-04-2024",
         118000, 100000, 18000, 0, 0, 0, ""],
        ["29AABCU9603R1ZN", "XYZ Traders",         "XYZ-24-0045",  "10-04-2024",
         59000,  50000,  0,    4500, 4500, 0, ""],
        ["07AABCU9603R1ZN", "Delhi Supplies Co",   "DS/APR/22",    "15-04-2024",
         23600,  20000,  3600, 0,    0,    0, ""],
        ["19AABCU9603R1ZN", "PQR Enterprises",     "PQR-0099",     "20-04-2024",
         141600, 120000, 21600, 0,   0,    0, ""],
        ["33AABCU9603R1ZN", "Chennai Parts Ltd",   "CPL/24/567",   "25-04-2024",
         35400,  30000,  0,    2700, 2700, 0, ""],
    ]
    for r_idx, row in enumerate(samples, 4):
        for c_idx, val in enumerate(row, 1):
            sample(ws, r_idx, c_idx, val)

    widths = [20, 22, 16, 14, 14, 14, 10, 10, 10, 8, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
    wb.save(path)
    print(f"Created: {path}")

def make_sales_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "SALES REGISTER – Sample Template for GST Reconciliation Tool"
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "Delete Row 1 & 2 before loading. Keep headers in Row 1 only. Date: DD-MM-YYYY"
    c.fill = PatternFill("solid", fgColor=MED)
    c.font = Font(bold=True, color=WHITE, size=9)
    c.alignment = Alignment(horizontal="center")

    headers = ["GSTIN", "Customer Name", "Invoice No", "Invoice Date",
               "Invoice Value", "Taxable Value", "IGST", "CGST", "SGST", "CESS",
               "Remarks"]
    for col, h in enumerate(headers, 1):
        hdr(ws, 3, col, h)

    samples = [
        ["27BBCDE1234F1Z5", "Raj Enterprises",     "SAL/24/001", "03-04-2024",
         118000, 100000, 18000, 0,    0,    0, ""],
        ["29BBCDE1234F1Z5", "Mysore Motors",        "SAL/24/002", "08-04-2024",
         59000,  50000,  0,    4500, 4500, 0, ""],
        ["07BBCDE1234F1Z5", "Kapoor & Sons",        "SAL/24/003", "12-04-2024",
         23600,  20000,  3600, 0,    0,    0, ""],
        ["19BBCDE1234F1Z5", "Kolkata Distributors","SAL/24/004", "18-04-2024",
         141600, 120000, 21600, 0,   0,    0, ""],
        ["33BBCDE1234F1Z5", "Southern Traders",    "SAL/24/005", "22-04-2024",
         35400,  30000,  0,    2700, 2700, 0, ""],
    ]
    for r_idx, row in enumerate(samples, 4):
        for c_idx, val in enumerate(row, 1):
            sample(ws, r_idx, c_idx, val)

    widths = [20, 22, 14, 14, 14, 14, 10, 10, 10, 8, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
    wb.save(path)
    print(f"Created: {path}")

def make_gstr3b_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-3B Data"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "GSTR-3B Summary – Template for 3B Comparison"
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Head", "Taxable Value", "IGST", "CGST", "SGST", "CESS"]
    for col, h in enumerate(headers, 1):
        hdr(ws, 2, col, h)

    rows = [
        ["3.1(a) Outward Taxable Supplies",     500000, 45000, 22500, 22500, 0],
        ["3.1(b) Outward Zero Rated",            100000, 0,     0,     0,     0],
        ["3.1(c) Other Outward Supplies",        20000,  3600,  0,     0,     0],
        ["4(A)(1) Import of Goods",              50000,  9000,  0,     0,     0],
        ["4(A)(5) All other ITC",                200000, 18000, 9000,  9000,  0],
    ]
    for r_idx, row in enumerate(rows, 3):
        for c_idx, val in enumerate(row, 1):
            sample(ws, r_idx, c_idx, val)

    for col, w in enumerate([35, 16, 12, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(path)
    print(f"Created: {path}")

def make_demo_json(path):
    """Create a minimal GSTR-2B JSON demo file."""
    import json
    demo = {
        "data": {
            "docdata": {
                "b2b": [
                    {
                        "ctin": "27AABCU9603R1ZN",
                        "trdnm": "ABC Pvt Ltd",
                        "inv": [
                            {
                                "inum": "INV/2024/001",
                                "idt": "05-04-2024",
                                "val": 118000,
                                "pos": "27",
                                "rev": "N",
                                "itcavl": "Y",
                                "items": [
                                    {
                                        "txval": 100000,
                                        "igst": 18000,
                                        "cgst": 0,
                                        "sgst": 0,
                                        "cess": 0
                                    }
                                ]
                            }
                        ]
                    },
                    {
                        "ctin": "29AABCU9603R1ZN",
                        "trdnm": "XYZ Traders",
                        "inv": [
                            {
                                "inum": "XYZ-24-0045",
                                "idt": "10-04-2024",
                                "val": 59000,
                                "pos": "29",
                                "rev": "N",
                                "itcavl": "Y",
                                "items": [
                                    {
                                        "txval": 50000,
                                        "igst": 0,
                                        "cgst": 4500,
                                        "sgst": 4500,
                                        "cess": 0
                                    }
                                ]
                            }
                        ]
                    },
                    {
                        "ctin": "07AABCU9603R1ZN",
                        "trdnm": "Delhi Supplies Co",
                        "inv": [
                            {
                                "inum": "DS/APR/22",
                                "idt": "15-04-2024",
                                "val": 23600,
                                "pos": "07",
                                "rev": "N",
                                "itcavl": "Y",
                                "items": [
                                    {
                                        "txval": 20000,
                                        "igst": 3600,
                                        "cgst": 0,
                                        "sgst": 0,
                                        "cess": 0
                                    }
                                ]
                            }
                        ]
                    },
                    {
                        "ctin": "40AABCU9603R1ZN",
                        "trdnm": "New Supplier (Portal Only)",
                        "inv": [
                            {
                                "inum": "NEW/001",
                                "idt": "28-04-2024",
                                "val": 47200,
                                "pos": "40",
                                "rev": "N",
                                "itcavl": "Y",
                                "items": [
                                    {
                                        "txval": 40000,
                                        "igst": 7200,
                                        "cgst": 0,
                                        "sgst": 0,
                                        "cess": 0
                                    }
                                ]
                            }
                        ]
                    }
                ]
            }
        }
    }
    with open(path, "w") as f:
        json.dump(demo, f, indent=2)
    print(f"Created: {path}")

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    tmpl = os.path.join(base, "Templates")
    os.makedirs(tmpl, exist_ok=True)
    make_purchase_template(os.path.join(tmpl, "Purchase_Register_Template.xlsx"))
    make_sales_template(   os.path.join(tmpl, "Sales_Register_Template.xlsx"))
    make_gstr3b_template(  os.path.join(tmpl, "GSTR3B_Template.xlsx"))
    make_demo_json(        os.path.join(tmpl, "Demo_GSTR2B.json"))
    print("\nAll templates created in:", tmpl)
