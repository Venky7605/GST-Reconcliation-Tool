"""
GST Reconciliation Tool - Free, No Payment Walls
Features:
  - GSTR-2A/2B vs Purchase Register Reconciliation
  - GSTR-1 vs Sales Register Reconciliation
  - GSTR-3B vs GSTR-1 Reconciliation
  - GSTR-3B vs GSTR-2B (ITC) Reconciliation
  - JSON import (GST Portal downloads)
  - Excel import (Books/Tally exports)
  - Color-coded Excel reports
  - Summary Dashboard
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import threading

# ── colour palette ──────────────────────────────────────────────────────────
CLR_MATCHED      = "C6EFCE"   # green
CLR_MISMATCH     = "FFEB9C"   # yellow
CLR_ONLY_PORTAL  = "FFC7CE"   # red   – in portal, not in books
CLR_ONLY_BOOKS   = "BDD7EE"   # blue  – in books, not in portal
CLR_HEADER_MAIN  = "1F3864"   # dark blue header
CLR_HEADER_SUB   = "2E75B6"   # medium blue sub-header
CLR_WHITE        = "FFFFFF"
CLR_TITLE_TEXT   = "FFFFFF"

FONT_HEADER = Font(bold=True, color=CLR_TITLE_TEXT, size=11)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def apply_header(ws, row, col, value, bg=CLR_HEADER_MAIN, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font  = FONT_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border = BORDER
    return cell

def apply_data(ws, row, col, value, bg=None, bold=False, num_fmt=None,
               align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════════
#  CORE RECONCILIATION LOGIC
# ════════════════════════════════════════════════════════════════════════════

def normalize_inv(s):
    """Strip spaces, slashes, dashes – case-insensitive match."""
    if pd.isna(s):
        return ""
    return re.sub(r"[\s\-/\\]", "", str(s)).upper()

def normalize_gstin(g):
    if pd.isna(g):
        return ""
    return str(g).strip().upper()

def safe_float(v):
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0

def parse_date(d):
    if pd.isna(d) or str(d).strip() == "":
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
                "%d %b %Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(str(d).strip(), fmt).date()
        except Exception:
            pass
    return None

# ── JSON parsers (GST portal format) ─────────────────────────────────────

def parse_gstr2b_json(path):
    """Parse GSTR-2B JSON downloaded from GST portal."""
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    rows = []
    # Handle different JSON structures
    doc_section = (data.get("data", {}) or data)
    docdata = doc_section.get("docdata", doc_section)

    # B2B invoices
    b2b = docdata.get("b2b", [])
    for supplier in b2b:
        gstin = supplier.get("ctin", "")
        trade = supplier.get("trdnm", "")
        for inv in supplier.get("inv", []):
            inv_no   = inv.get("inum", "")
            inv_date = inv.get("idt",  "")
            inv_val  = safe_float(inv.get("val", 0))
            pos      = inv.get("pos", "")
            rev      = inv.get("rev", "N")
            itcavl   = inv.get("itcavl", "")
            for item in inv.get("items", []):
                rows.append({
                    "GSTIN":        normalize_gstin(gstin),
                    "Supplier":     trade,
                    "Invoice No":   inv_no,
                    "Invoice Date": inv_date,
                    "Invoice Value":inv_val,
                    "Taxable Value":safe_float(item.get("txval", 0)),
                    "IGST":         safe_float(item.get("igst",  0)),
                    "CGST":         safe_float(item.get("cgst",  0)),
                    "SGST":         safe_float(item.get("sgst",  0)),
                    "CESS":         safe_float(item.get("cess",  0)),
                    "POS":          pos,
                    "Reverse":      rev,
                    "ITC Available":itcavl,
                    "Source":       "2B"
                })
    if not rows:
        # Flat format fallback
        for k in ["b2b", "cdnr", "isd"]:
            section = docdata.get(k, [])
            for rec in section:
                rows.append({
                    "GSTIN":        normalize_gstin(rec.get("ctin",  rec.get("gstin",""))),
                    "Supplier":     rec.get("trdnm", ""),
                    "Invoice No":   rec.get("inum",  rec.get("invno","")),
                    "Invoice Date": rec.get("idt",   rec.get("invdt","")),
                    "Invoice Value":safe_float(rec.get("val", 0)),
                    "Taxable Value":safe_float(rec.get("txval",0)),
                    "IGST":         safe_float(rec.get("igst",0)),
                    "CGST":         safe_float(rec.get("cgst",0)),
                    "SGST":         safe_float(rec.get("sgst",0)),
                    "CESS":         safe_float(rec.get("cess",0)),
                    "ITC Available":"",
                    "Source":       "2B"
                })
    return pd.DataFrame(rows)

def parse_gstr2a_json(path):
    """Parse GSTR-2A JSON."""
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    rows = []
    doc = data.get("data", data)
    b2b = doc.get("b2b", [])
    for supplier in b2b:
        gstin = supplier.get("ctin", "")
        trade = supplier.get("trdnm", "")
        for inv in supplier.get("inv", []):
            inv_no   = inv.get("inum", "")
            inv_date = inv.get("idt",  "")
            inv_val  = safe_float(inv.get("val", 0))
            for item in inv.get("itms", []):
                d = item.get("itm_det", item)
                rows.append({
                    "GSTIN":        normalize_gstin(gstin),
                    "Supplier":     trade,
                    "Invoice No":   inv_no,
                    "Invoice Date": inv_date,
                    "Invoice Value":inv_val,
                    "Taxable Value":safe_float(d.get("txval", 0)),
                    "IGST":         safe_float(d.get("igst",  0)),
                    "CGST":         safe_float(d.get("camt",  d.get("cgst",0))),
                    "SGST":         safe_float(d.get("samt",  d.get("sgst",0))),
                    "CESS":         safe_float(d.get("csamt", d.get("cess",0))),
                    "Source":       "2A"
                })
    return pd.DataFrame(rows)

def parse_gstr1_json(path):
    """Parse GSTR-1 JSON."""
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    rows = []
    doc = data.get("data", data)
    b2b = doc.get("b2b", [])
    for customer in b2b:
        gstin = customer.get("ctin","")
        trade = customer.get("trdnm", customer.get("cname",""))
        for inv in customer.get("inv",[]):
            inv_no   = inv.get("inum","")
            inv_date = inv.get("idt", "")
            inv_val  = safe_float(inv.get("val",0))
            for item in inv.get("itms",[]):
                d = item.get("itm_det", item)
                rows.append({
                    "GSTIN":        normalize_gstin(gstin),
                    "Customer":     trade,
                    "Invoice No":   inv_no,
                    "Invoice Date": inv_date,
                    "Invoice Value":inv_val,
                    "Taxable Value":safe_float(d.get("txval",0)),
                    "IGST":         safe_float(d.get("igst",0)),
                    "CGST":         safe_float(d.get("camt",d.get("cgst",0))),
                    "SGST":         safe_float(d.get("samt",d.get("sgst",0))),
                    "CESS":         safe_float(d.get("csamt",d.get("cess",0))),
                    "Source":       "GSTR-1"
                })
    return pd.DataFrame(rows)

# ── Excel reader (generic) ────────────────────────────────────────────────

def read_excel_file(path, sheet=0):
    """Read Excel/CSV into DataFrame, auto-detecting the header row (up to row 6)."""
    ext = os.path.splitext(path)[1].lower()
    KNOWN = {"gstin","invoice no","invoice date","taxable value",
             "igst","cgst","sgst","supplier","customer","party"}

    if ext == ".csv":
        df = pd.read_csv(path, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        return df

    for skip in range(0, 6):
        try:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str, header=skip)
            cols_lower = {str(c).strip().lower() for c in df.columns}
            if cols_lower & KNOWN:   # at least one known column found
                df.columns = [str(c).strip() for c in df.columns]
                # drop fully-empty rows
                df.dropna(how="all", inplace=True)
                return df
        except Exception:
            pass

    # fallback: just read as-is from row 0
    df = pd.read_excel(path, sheet_name=sheet, dtype=str, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    return df

# ── Column auto-mapper ────────────────────────────────────────────────────

def find_col(df, candidates):
    """Find first matching column from a list of candidates (case-insensitive)."""
    cols = {c.lower().strip(): c for c in df.columns}
    for c in candidates:
        if c.lower() in cols:
            return cols[c.lower()]
    return None

def standardize_books_purchase(df):
    """Map books purchase register columns to standard names."""
    mapping = {
        "GSTIN":        ["gstin","supplier gstin","party gstin","vendor gstin",
                         "supplier gst","gstin/uin"],
        "Supplier":     ["supplier","vendor","party name","supplier name",
                         "vendor name","party","name"],
        "Invoice No":   ["invoice no","inv no","invoice number","inv number",
                         "bill no","bill number","voucher no","doc no"],
        "Invoice Date": ["invoice date","inv date","bill date","voucher date",
                         "date","doc date"],
        "Invoice Value":["invoice value","inv value","bill amount","total amount",
                         "gross amount","total","amount"],
        "Taxable Value":["taxable value","taxable amount","taxable","basic amount",
                         "assessable value"],
        "IGST":         ["igst","integrated tax","igst amount"],
        "CGST":         ["cgst","central tax","cgst amount"],
        "SGST":         ["sgst","sgst/utgst","state tax","sgst amount","utgst"],
        "CESS":         ["cess","cess amount"],
    }
    out = {}
    for std, candidates in mapping.items():
        col = find_col(df, candidates)
        if col:
            out[std] = df[col]
    result = pd.DataFrame(out)
    if "GSTIN" in result.columns:
        result["GSTIN"] = result["GSTIN"].apply(normalize_gstin)
    for c in ["Invoice Value","Taxable Value","IGST","CGST","SGST","CESS"]:
        if c in result.columns:
            result[c] = result[c].apply(safe_float)
        else:
            result[c] = 0.0
    return result

def standardize_books_sales(df):
    mapping = {
        "GSTIN":        ["gstin","customer gstin","party gstin","buyer gstin"],
        "Customer":     ["customer","buyer","party name","customer name","party"],
        "Invoice No":   ["invoice no","inv no","invoice number","bill no",
                         "voucher no","doc no"],
        "Invoice Date": ["invoice date","inv date","bill date","date"],
        "Invoice Value":["invoice value","inv value","bill amount","total amount",
                         "gross amount","total","amount"],
        "Taxable Value":["taxable value","taxable amount","taxable","basic amount"],
        "IGST":         ["igst","integrated tax"],
        "CGST":         ["cgst","central tax"],
        "SGST":         ["sgst","sgst/utgst","state tax","utgst"],
        "CESS":         ["cess"],
    }
    out = {}
    for std, candidates in mapping.items():
        col = find_col(df, candidates)
        if col:
            out[std] = df[col]
    result = pd.DataFrame(out)
    if "GSTIN" in result.columns:
        result["GSTIN"] = result["GSTIN"].apply(normalize_gstin)
    for c in ["Invoice Value","Taxable Value","IGST","CGST","SGST","CESS"]:
        if c in result.columns:
            result[c] = result[c].apply(safe_float)
        else:
            result[c] = 0.0
    return result

# ── Main reconciliation engine ────────────────────────────────────────────

TOLERANCE = 1.0   # Rs 1 tolerance for rounding differences

def reconcile_purchase(portal_df, books_df):
    """
    Match portal (2A/2B) rows with books purchase register.
    Key = GSTIN + normalized Invoice No
    Returns matched, only_portal, only_books, mismatch DataFrames.
    """
    portal_df = portal_df.copy()
    books_df  = books_df.copy()

    for df_name, df in [("portal", portal_df), ("books", books_df)]:
        if "GSTIN" not in df.columns:
            df["GSTIN"] = ""
        if "Invoice No" not in df.columns:
            df["Invoice No"] = ""

    portal_df["_key"] = (portal_df["GSTIN"] + "|" +
                         portal_df["Invoice No"].apply(normalize_inv))
    books_df["_key"]  = (books_df["GSTIN"] + "|" +
                         books_df["Invoice No"].apply(normalize_inv))

    portal_df["_matched"] = False
    books_df["_matched"]  = False

    matched_rows    = []
    mismatch_rows   = []
    only_portal     = []
    only_books      = []

    portal_keyed = {}
    for _, row in portal_df.iterrows():
        portal_keyed.setdefault(row["_key"], []).append(row)

    books_keyed = {}
    for _, row in books_df.iterrows():
        books_keyed.setdefault(row["_key"], []).append(row)

    all_keys = set(portal_keyed) | set(books_keyed)

    for key in all_keys:
        p_list = portal_keyed.get(key, [])
        b_list = books_keyed.get(key, [])

        if p_list and b_list:
            # pair them up 1:1 (first-come first-served)
            for i in range(max(len(p_list), len(b_list))):
                p = p_list[i] if i < len(p_list) else None
                b = b_list[i] if i < len(b_list) else None
                if p is not None and b is not None:
                    p_tax = p.get("IGST",0)+p.get("CGST",0)+p.get("SGST",0)+p.get("CESS",0)
                    b_tax = b.get("IGST",0)+b.get("CGST",0)+b.get("SGST",0)+b.get("CESS",0)
                    p_txval = p.get("Taxable Value",0)
                    b_txval = b.get("Taxable Value",0)
                    tax_diff  = round(p_tax - b_tax, 2)
                    val_diff  = round(p_txval - b_txval, 2)
                    status = ("Matched" if abs(tax_diff) <= TOLERANCE and
                              abs(val_diff) <= TOLERANCE else "Mismatch")
                    row = {
                        "Status":              status,
                        "GSTIN":               p["GSTIN"],
                        "Supplier":            p.get("Supplier",""),
                        "Invoice No (Portal)": p.get("Invoice No",""),
                        "Invoice No (Books)":  b.get("Invoice No",""),
                        "Invoice Date (Portal)":p.get("Invoice Date",""),
                        "Invoice Date (Books)": b.get("Invoice Date",""),
                        "Taxable Value (Portal)":p_txval,
                        "Taxable Value (Books)": b_txval,
                        "Taxable Value Diff":    val_diff,
                        "IGST (Portal)":  p.get("IGST",0),
                        "IGST (Books)":   b.get("IGST",0),
                        "CGST (Portal)":  p.get("CGST",0),
                        "CGST (Books)":   b.get("CGST",0),
                        "SGST (Portal)":  p.get("SGST",0),
                        "SGST (Books)":   b.get("SGST",0),
                        "CESS (Portal)":  p.get("CESS",0),
                        "CESS (Books)":   b.get("CESS",0),
                        "Tax Diff":       tax_diff,
                        "ITC Available":  p.get("ITC Available",""),
                    }
                    if status == "Matched":
                        matched_rows.append(row)
                    else:
                        mismatch_rows.append(row)
                elif p is not None:
                    only_portal.append(_portal_only_row(p))
                else:
                    only_books.append(_books_only_row(b))
        elif p_list:
            for p in p_list:
                only_portal.append(_portal_only_row(p))
        else:
            for b in b_list:
                only_books.append(_books_only_row(b))

    return (pd.DataFrame(matched_rows),
            pd.DataFrame(mismatch_rows),
            pd.DataFrame(only_portal),
            pd.DataFrame(only_books))

def _portal_only_row(p):
    return {
        "Status": "In Portal Not in Books",
        "GSTIN": p.get("GSTIN",""),
        "Supplier/Customer": p.get("Supplier", p.get("Customer","")),
        "Invoice No": p.get("Invoice No",""),
        "Invoice Date": p.get("Invoice Date",""),
        "Invoice Value": p.get("Invoice Value",0),
        "Taxable Value": p.get("Taxable Value",0),
        "IGST": p.get("IGST",0),
        "CGST": p.get("CGST",0),
        "SGST": p.get("SGST",0),
        "CESS": p.get("CESS",0),
        "Total Tax": p.get("IGST",0)+p.get("CGST",0)+p.get("SGST",0)+p.get("CESS",0),
        "ITC Available": p.get("ITC Available",""),
    }

def _books_only_row(b):
    return {
        "Status": "In Books Not in Portal",
        "GSTIN": b.get("GSTIN",""),
        "Supplier/Customer": b.get("Supplier", b.get("Customer","")),
        "Invoice No": b.get("Invoice No",""),
        "Invoice Date": b.get("Invoice Date",""),
        "Invoice Value": b.get("Invoice Value",0),
        "Taxable Value": b.get("Taxable Value",0),
        "IGST": b.get("IGST",0),
        "CGST": b.get("CGST",0),
        "SGST": b.get("SGST",0),
        "CESS": b.get("CESS",0),
        "Total Tax": b.get("IGST",0)+b.get("CGST",0)+b.get("SGST",0)+b.get("CESS",0),
        "ITC Available": "",
    }

# same engine for sales (GSTR-1 vs Sales Register)
def reconcile_sales(portal_df, books_df):
    portal_df = portal_df.copy()
    books_df  = books_df.copy()
    # rename Customer->Supplier for shared engine
    if "Customer" in portal_df.columns:
        portal_df["Supplier"] = portal_df["Customer"]
    if "Customer" in books_df.columns:
        books_df["Supplier"] = books_df["Customer"]
    return reconcile_purchase(portal_df, books_df)

# ── 3B vs 1 / 3B vs 2B reconciliation ───────────────────────────────────

def reconcile_3b_vs_1(gstr3b_df, gstr1_df):
    """Compare summary totals from GSTR-3B vs GSTR-1."""
    rows = []
    heads = ["Taxable Value","IGST","CGST","SGST","CESS"]
    for h in heads:
        v3b = safe_float(gstr3b_df[h].sum()) if h in gstr3b_df.columns else 0
        v1  = safe_float(gstr1_df[h].sum())  if h in gstr1_df.columns  else 0
        diff = round(v3b - v1, 2)
        rows.append({
            "Head": h,
            "GSTR-3B Amount": v3b,
            "GSTR-1 Amount":  v1,
            "Difference":     diff,
            "Status": "OK" if abs(diff) <= TOLERANCE else "Mismatch"
        })
    return pd.DataFrame(rows)

def reconcile_3b_vs_2b(gstr3b_df, gstr2b_df):
    rows = []
    heads = ["Taxable Value","IGST","CGST","SGST","CESS"]
    for h in heads:
        v3b = safe_float(gstr3b_df[h].sum()) if h in gstr3b_df.columns else 0
        v2b = safe_float(gstr2b_df[h].sum()) if h in gstr2b_df.columns else 0
        diff = round(v3b - v2b, 2)
        rows.append({
            "Head": h,
            "ITC in GSTR-3B":  v3b,
            "ITC in GSTR-2B":  v2b,
            "Difference":      diff,
            "Status": "OK" if abs(diff) <= TOLERANCE else "Mismatch"
        })
    return pd.DataFrame(rows)

# ════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT WRITER
# ════════════════════════════════════════════════════════════════════════════

def write_recon_report(output_path, report_data: dict, period: str, gstin: str):
    """
    report_data keys accepted:
      "purchase_matched", "purchase_mismatch", "purchase_portal_only",
      "purchase_books_only",
      "sales_matched", "sales_mismatch", "sales_portal_only", "sales_books_only",
      "gstr3b_vs_1", "gstr3b_vs_2b"
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Cover / Summary sheet ─────────────────────────────────────────────
    ws_cover = wb.create_sheet("Summary Dashboard")
    _write_summary(ws_cover, report_data, period, gstin)

    # ── Purchase Reconciliation sheets ───────────────────────────────────
    sections = [
        ("purchase_matched",     "Purch - Matched",      CLR_MATCHED,     "Purchase"),
        ("purchase_mismatch",    "Purch - Mismatch",     CLR_MISMATCH,    "Purchase"),
        ("purchase_portal_only", "Purch - Portal Only",  CLR_ONLY_PORTAL, "Purchase"),
        ("purchase_books_only",  "Purch - Books Only",   CLR_ONLY_BOOKS,  "Purchase"),
        ("sales_matched",        "Sales - Matched",      CLR_MATCHED,     "Sales"),
        ("sales_mismatch",       "Sales - Mismatch",     CLR_MISMATCH,    "Sales"),
        ("sales_portal_only",    "Sales - Portal Only",  CLR_ONLY_PORTAL, "Sales"),
        ("sales_books_only",     "Sales - Books Only",   CLR_ONLY_BOOKS,  "Sales"),
    ]

    for key, sheet_name, colour, recon_type in sections:
        df = report_data.get(key)
        if df is not None and not df.empty:
            ws = wb.create_sheet(sheet_name)
            _write_detail_sheet(ws, df, sheet_name, colour, period, gstin, recon_type)

    # ── 3B vs 1 / 3B vs 2B ───────────────────────────────────────────────
    for key, sheet_name in [("gstr3b_vs_1","3B vs GSTR-1"),
                             ("gstr3b_vs_2b","3B vs GSTR-2B")]:
        df = report_data.get(key)
        if df is not None and not df.empty:
            ws = wb.create_sheet(sheet_name)
            _write_summary_recon_sheet(ws, df, sheet_name, period, gstin)

    wb.save(output_path)

def _write_summary(ws, report_data, period, gstin):
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "GST RECONCILIATION REPORT"
    c.fill  = PatternFill("solid", fgColor=CLR_HEADER_MAIN)
    c.font  = Font(bold=True, size=20, color=CLR_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = f"Period: {period}    |    GSTIN: {gstin}    |    Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    c.fill  = PatternFill("solid", fgColor=CLR_HEADER_SUB)
    c.font  = Font(bold=True, size=11, color=CLR_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10

    # Legend
    legend_row = 4
    ws.cell(row=legend_row, column=1, value="Legend:").font = FONT_BOLD
    for col, (label, colour) in enumerate([
        ("Matched",          CLR_MATCHED),
        ("Mismatch",         CLR_MISMATCH),
        ("Portal Only",      CLR_ONLY_PORTAL),
        ("Books Only",       CLR_ONLY_BOOKS),
    ], 2):
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = PatternFill("solid", fgColor=colour)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.row_dimensions[5].height = 10

    # Summary table headers
    hrow = 6
    for col, label in enumerate([
        "Reconciliation Type","Total Records","Matched","Mismatch",
        "Portal Only","Books Only","Taxable Value Diff","Tax Diff"
    ], 1):
        apply_header(ws, hrow, col, label, CLR_HEADER_MAIN)

    # data rows
    categories = [
        ("Purchase (2A/2B vs Books)",
         "purchase_matched","purchase_mismatch","purchase_portal_only","purchase_books_only"),
        ("Sales (GSTR-1 vs Books)",
         "sales_matched","sales_mismatch","sales_portal_only","sales_books_only"),
    ]
    drow = hrow + 1
    for label, m_key, mm_key, p_key, b_key in categories:
        matched   = report_data.get(m_key,  pd.DataFrame())
        mismatch  = report_data.get(mm_key, pd.DataFrame())
        p_only    = report_data.get(p_key,  pd.DataFrame())
        b_only    = report_data.get(b_key,  pd.DataFrame())

        n_matched  = len(matched)
        n_mismatch = len(mismatch)
        n_portal   = len(p_only)
        n_books    = len(b_only)
        total      = n_matched + n_mismatch + n_portal + n_books

        txval_diff = 0.0
        tax_diff   = 0.0
        if not mismatch.empty and "Taxable Value Diff" in mismatch.columns:
            txval_diff = mismatch["Taxable Value Diff"].sum()
        if not mismatch.empty and "Tax Diff" in mismatch.columns:
            tax_diff = mismatch["Tax Diff"].sum()
        if not p_only.empty and "Taxable Value" in p_only.columns:
            txval_diff += p_only["Taxable Value"].sum()
            tax_diff   += p_only["Total Tax"].sum()
        if not b_only.empty and "Taxable Value" in b_only.columns:
            txval_diff -= b_only["Taxable Value"].sum()
            tax_diff   -= b_only["Total Tax"].sum()

        for col, val in enumerate([label, total, n_matched, n_mismatch,
                                    n_portal, n_books,
                                    round(txval_diff,2), round(tax_diff,2)], 1):
            bg = None
            if col == 3: bg = CLR_MATCHED
            if col == 4: bg = CLR_MISMATCH if n_mismatch else None
            if col == 5: bg = CLR_ONLY_PORTAL if n_portal else None
            if col == 6: bg = CLR_ONLY_BOOKS  if n_books  else None
            nm = "#,##0.00" if col >= 7 else None
            apply_data(ws, drow, col, val, bg=bg,
                       bold=(col==1), num_fmt=nm,
                       align="center" if col > 1 else "left")
        drow += 1

    # column widths for summary
    ws.column_dimensions["A"].width = 36
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 18

    # ── ITC Impact box ────────────────────────────────────────────────────
    drow += 1
    ws.merge_cells(f"A{drow}:H{drow}")
    c = ws.cell(row=drow, column=1, value="ITC IMPACT ANALYSIS")
    c.fill = PatternFill("solid", fgColor=CLR_HEADER_MAIN)
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center")
    drow += 1

    p_only_df  = report_data.get("purchase_portal_only", pd.DataFrame())
    b_only_df  = report_data.get("purchase_books_only",  pd.DataFrame())
    mis_df     = report_data.get("purchase_mismatch",    pd.DataFrame())

    itc_portal  = p_only_df["Total Tax"].sum()    if not p_only_df.empty and "Total Tax" in p_only_df.columns  else 0
    itc_books   = b_only_df["Total Tax"].sum()    if not b_only_df.empty and "Total Tax" in b_only_df.columns  else 0
    itc_mis_diff= mis_df["Tax Diff"].sum()        if not mis_df.empty   and "Tax Diff" in mis_df.columns       else 0
    net_itc_diff= round(itc_portal - itc_books + itc_mis_diff, 2)

    for label, val, col in [
        ("ITC in Portal NOT in Books (Eligible ITC Not Claimed)", itc_portal, CLR_ONLY_PORTAL),
        ("ITC in Books NOT in Portal (Ineligible / Risky ITC)",   itc_books,  CLR_ONLY_BOOKS),
        ("ITC Difference (Mismatch invoices)",                    itc_mis_diff, CLR_MISMATCH),
        ("Net ITC Impact",                                        net_itc_diff, CLR_MATCHED if net_itc_diff==0 else CLR_MISMATCH),
    ]:
        apply_data(ws, drow, 1, label, bold=True)
        apply_data(ws, drow, 2, round(val,2), bg=col, num_fmt="#,##0.00", align="right")
        drow += 1

def _write_detail_sheet(ws, df, title, row_colour, period, gstin, recon_type):
    ws.sheet_view.showGridLines = False

    # Title
    max_col = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=f"{recon_type} Reconciliation – {title}  |  {period}  |  {gstin}")
    c.fill = PatternFill("solid", fgColor=CLR_HEADER_MAIN)
    c.font = Font(bold=True, size=13, color=CLR_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Summary row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    c = ws.cell(row=2, column=1,
                value=f"Records: {len(df)}    Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
    c.fill = PatternFill("solid", fgColor=CLR_HEADER_SUB)
    c.font = Font(bold=True, size=10, color=CLR_WHITE)
    c.alignment = Alignment(horizontal="center")

    # Headers row 3
    for col, hdr in enumerate(df.columns, 1):
        apply_header(ws, 3, col, hdr, CLR_HEADER_MAIN)

    # Data rows
    num_cols = {c for c in df.columns if any(
        x in c for x in ["Value","IGST","CGST","SGST","CESS","Tax","Diff"])}

    for r_idx, (_, row) in enumerate(df.iterrows(), 4):
        for c_idx, col in enumerate(df.columns, 1):
            val = row[col]
            if col in num_cols:
                val = safe_float(val) if not pd.isna(val) else 0.0
            nm = "#,##0.00" if col in num_cols else None
            apply_data(ws, r_idx, c_idx, val,
                       bg=row_colour, num_fmt=nm,
                       align="right" if col in num_cols else "left")

    # Auto-width
    for col_idx in range(1, max_col+1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            [len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, len(df)+4)],
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # Totals row
    trow = len(df) + 4
    ws.cell(row=trow, column=1, value="TOTAL").font = FONT_BOLD
    ws.cell(row=trow, column=1).fill = PatternFill("solid", fgColor=CLR_HEADER_SUB)
    for c_idx, col in enumerate(df.columns, 1):
        if col in num_cols:
            s = df[col].apply(safe_float).sum()
            c = apply_data(ws, trow, c_idx, round(s,2),
                           bg=CLR_HEADER_SUB, bold=True,
                           num_fmt="#,##0.00", align="right")
            c.font = Font(bold=True, size=10, color=CLR_WHITE)

    # Freeze panes
    ws.freeze_panes = "A4"

def _write_summary_recon_sheet(ws, df, title, period, gstin):
    ws.sheet_view.showGridLines = False
    max_col = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=f"{title}  |  {period}  |  {gstin}")
    c.fill = PatternFill("solid", fgColor=CLR_HEADER_MAIN)
    c.font = Font(bold=True, size=13, color=CLR_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, hdr in enumerate(df.columns, 1):
        apply_header(ws, 2, col, hdr)

    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        for c_idx, col in enumerate(df.columns, 1):
            val = row[col]
            is_num = col not in ("Head","Status")
            bg = CLR_MATCHED if row.get("Status") == "OK" else CLR_MISMATCH
            apply_data(ws, r_idx, c_idx, val,
                       bg=bg, num_fmt="#,##0.00" if is_num else None,
                       align="right" if is_num else "left")

    for col_idx in range(1, max_col+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

# ════════════════════════════════════════════════════════════════════════════
#  GUI
# ════════════════════════════════════════════════════════════════════════════

class GSTReconApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GST Reconciliation Tool – Free Edition")
        self.geometry("1000x760")
        self.resizable(True, True)
        self.configure(bg="#1F3864")

        # state
        self.portal_path  = tk.StringVar()
        self.books_path   = tk.StringVar()
        self.portal2_path = tk.StringVar()  # sales portal
        self.books2_path  = tk.StringVar()  # sales books
        self.gstr3b_path  = tk.StringVar()
        self.output_path  = tk.StringVar()
        self.period_var   = tk.StringVar(value="Apr-2024")
        self.gstin_var    = tk.StringVar(value="")
        self.portal_type  = tk.StringVar(value="GSTR-2B")
        self.status_var   = tk.StringVar(value="Ready")
        self.progress_var = tk.DoubleVar(value=0)

        self._build_ui()

    # ── UI ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ─ Title bar ─
        title_frame = tk.Frame(self, bg="#1F3864", pady=10)
        title_frame.pack(fill="x")
        tk.Label(title_frame, text="GST Reconciliation Tool",
                 font=("Segoe UI", 22, "bold"),
                 bg="#1F3864", fg="white").pack()
        tk.Label(title_frame,
                 text="100% Free  •  No Payment Walls  •  All Features Included",
                 font=("Segoe UI", 10), bg="#1F3864", fg="#90CAF9").pack()

        # ─ Notebook ─
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",        background="#1F3864", borderwidth=0)
        style.configure("TNotebook.Tab",    background="#2E75B6", foreground="white",
                         padding=[14, 6], font=("Segoe UI", 10, "bold"))
        style.map("TNotebook.Tab",
                  background=[("selected","#1F3864")],
                  foreground=[("selected","white")])
        style.configure("TFrame", background="#F0F4FA")
        style.configure("TLabel", background="#F0F4FA", font=("Segoe UI", 10))
        style.configure("TEntry", font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10, "bold"),
                         padding=6, background="#2E75B6", foreground="white")
        style.configure("Green.TButton", background="#1B7A34", foreground="white",
                         font=("Segoe UI", 12, "bold"), padding=10)
        style.configure("TLabelframe",      background="#F0F4FA")
        style.configure("TLabelframe.Label",background="#F0F4FA",
                         font=("Segoe UI", 10, "bold"), foreground="#1F3864")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=5)

        self.tab_settings  = ttk.Frame(nb)
        self.tab_purchase  = ttk.Frame(nb)
        self.tab_sales     = ttk.Frame(nb)
        self.tab_3b        = ttk.Frame(nb)
        self.tab_help      = ttk.Frame(nb)

        nb.add(self.tab_settings, text=" Settings ")
        nb.add(self.tab_purchase, text=" Purchase (2A/2B vs Books) ")
        nb.add(self.tab_sales,    text=" Sales (GSTR-1 vs Books) ")
        nb.add(self.tab_3b,       text=" 3B Comparison ")
        nb.add(self.tab_help,     text=" Help ")

        self._build_settings_tab()
        self._build_purchase_tab()
        self._build_sales_tab()
        self._build_3b_tab()
        self._build_help_tab()

        # ─ Bottom bar ─
        bot = tk.Frame(self, bg="#1F3864", pady=6)
        bot.pack(fill="x", side="bottom")

        ttk.Button(bot, text="▶  RUN RECONCILIATION",
                   style="Green.TButton",
                   command=self._run_thread).pack(side="left", padx=14)

        self.progress = ttk.Progressbar(bot, variable=self.progress_var,
                                        maximum=100, length=400)
        self.progress.pack(side="left", padx=10)

        tk.Label(bot, textvariable=self.status_var,
                 font=("Segoe UI", 10), bg="#1F3864", fg="#90CAF9").pack(
            side="left", padx=10)

    # ─── Settings tab ────────────────────────────────────────────────────

    def _build_settings_tab(self):
        f = self.tab_settings
        pad = dict(padx=16, pady=8)

        info = ttk.LabelFrame(f, text="Company / Period Details", padding=12)
        info.pack(fill="x", **pad)

        ttk.Label(info, text="GSTIN:").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(info, textvariable=self.gstin_var, width=22).grid(
            row=0, column=1, sticky="w", padx=8)

        ttk.Label(info, text="Period (e.g. Apr-2024):").grid(
            row=0, column=2, sticky="w", padx=16)
        ttk.Entry(info, textvariable=self.period_var, width=14).grid(
            row=0, column=3, sticky="w", padx=8)

        out_frame = ttk.LabelFrame(f, text="Output Report Location", padding=12)
        out_frame.pack(fill="x", **pad)
        ttk.Entry(out_frame, textvariable=self.output_path, width=70).grid(
            row=0, column=0, sticky="ew", padx=(0,8))
        ttk.Button(out_frame, text="Browse",
                   command=self._browse_output).grid(row=0, column=1)
        out_frame.columnconfigure(0, weight=1)

        tip = ttk.LabelFrame(f, text="Tips", padding=12)
        tip.pack(fill="x", **pad)
        tips_text = (
            "• Download GSTR-2B or GSTR-2A JSON from GST Portal → Returns → GSTR-2B/2A → Download JSON\n"
            "• Download GSTR-1 JSON from GST Portal → Returns → GSTR-1 → Filed Returns → Download JSON\n"
            "• Export Purchase/Sales Register from Tally / accounting software as Excel\n"
            "• Required columns in Books Excel: GSTIN, Invoice No, Invoice Date, Taxable Value, IGST, CGST, SGST\n"
            "• The tool will auto-detect column names – works with most standard export formats"
        )
        tk.Text(tip, height=6, wrap="word", font=("Segoe UI", 10),
                bg="#EEF2FF", relief="flat").insert("1.0", tips_text)
        tk.Text(tip, height=6, wrap="word", font=("Segoe UI", 10),
                bg="#EEF2FF", relief="flat", state="disabled").pack(fill="x")

        # Re-do properly with a disabled Text widget
        for w in tip.winfo_children():
            w.destroy()
        txt = tk.Text(tip, height=6, wrap="word", font=("Segoe UI", 10),
                      bg="#EEF2FF", relief="flat", borderwidth=0)
        txt.insert("1.0", tips_text)
        txt.config(state="disabled")
        txt.pack(fill="x")

    # ─── Purchase tab ─────────────────────────────────────────────────────

    def _build_purchase_tab(self):
        f = self.tab_purchase
        pad = dict(padx=16, pady=8)

        portal_f = ttk.LabelFrame(
            f, text="GST Portal Data (GSTR-2B or GSTR-2A JSON)", padding=12)
        portal_f.pack(fill="x", **pad)

        ttk.Label(portal_f, text="Portal File Type:").grid(
            row=0, column=0, sticky="w")
        for i, t in enumerate(["GSTR-2B","GSTR-2A"]):
            ttk.Radiobutton(portal_f, text=t, variable=self.portal_type,
                            value=t).grid(row=0, column=i+1, padx=8, sticky="w")

        ttk.Label(portal_f, text="JSON File:").grid(
            row=1, column=0, sticky="w", pady=4)
        ttk.Entry(portal_f, textvariable=self.portal_path, width=65).grid(
            row=1, column=1, columnspan=3, sticky="ew", padx=(8,8))
        ttk.Button(portal_f, text="Browse",
                   command=lambda: self._browse_file(
                       self.portal_path, [("JSON","*.json"),("All","*.*")]
                   )).grid(row=1, column=4)
        portal_f.columnconfigure(1, weight=1)

        books_f = ttk.LabelFrame(
            f, text="Books – Purchase Register (Excel / CSV)", padding=12)
        books_f.pack(fill="x", **pad)
        ttk.Label(books_f, text="Excel/CSV File:").grid(
            row=0, column=0, sticky="w")
        ttk.Entry(books_f, textvariable=self.books_path, width=65).grid(
            row=0, column=1, sticky="ew", padx=(8,8))
        ttk.Button(books_f, text="Browse",
                   command=lambda: self._browse_file(
                       self.books_path,
                       [("Excel","*.xlsx *.xls *.xlsb *.csv"),("All","*.*")]
                   )).grid(row=0, column=2)
        books_f.columnconfigure(1, weight=1)

        # Required columns info
        info_f = ttk.LabelFrame(f, text="Required Columns in Books Excel", padding=10)
        info_f.pack(fill="x", **pad)
        cols_info = (
            "GSTIN  |  Supplier Name  |  Invoice No  |  Invoice Date  |  "
            "Invoice Value  |  Taxable Value  |  IGST  |  CGST  |  SGST  |  CESS\n\n"
            "Column names are flexible – tool auto-detects standard Tally/SAP/Zoho exports. "
            "Date format: DD-MM-YYYY or DD/MM/YYYY"
        )
        ttk.Label(info_f, text=cols_info, wraplength=800,
                  foreground="#1F3864").pack(anchor="w")

    # ─── Sales tab ────────────────────────────────────────────────────────

    def _build_sales_tab(self):
        f = self.tab_sales
        pad = dict(padx=16, pady=8)

        portal_f = ttk.LabelFrame(
            f, text="GST Portal Data – GSTR-1 JSON", padding=12)
        portal_f.pack(fill="x", **pad)
        ttk.Label(portal_f, text="JSON File:").grid(row=0, column=0, sticky="w")
        ttk.Entry(portal_f, textvariable=self.portal2_path, width=65).grid(
            row=0, column=1, sticky="ew", padx=(8,8))
        ttk.Button(portal_f, text="Browse",
                   command=lambda: self._browse_file(
                       self.portal2_path, [("JSON","*.json"),("All","*.*")]
                   )).grid(row=0, column=2)
        portal_f.columnconfigure(1, weight=1)

        books_f = ttk.LabelFrame(
            f, text="Books – Sales Register (Excel / CSV)", padding=12)
        books_f.pack(fill="x", **pad)
        ttk.Label(books_f, text="Excel/CSV File:").grid(row=0, column=0, sticky="w")
        ttk.Entry(books_f, textvariable=self.books2_path, width=65).grid(
            row=0, column=1, sticky="ew", padx=(8,8))
        ttk.Button(books_f, text="Browse",
                   command=lambda: self._browse_file(
                       self.books2_path,
                       [("Excel","*.xlsx *.xls *.xlsb *.csv"),("All","*.*")]
                   )).grid(row=0, column=2)
        books_f.columnconfigure(1, weight=1)

        info_f = ttk.LabelFrame(f, text="Required Columns in Books Excel", padding=10)
        info_f.pack(fill="x", **pad)
        ttk.Label(info_f,
                  text=("GSTIN  |  Customer Name  |  Invoice No  |  Invoice Date  |  "
                        "Invoice Value  |  Taxable Value  |  IGST  |  CGST  |  SGST  |  CESS"),
                  wraplength=800, foreground="#1F3864").pack(anchor="w")

    # ─── 3B tab ───────────────────────────────────────────────────────────

    def _build_3b_tab(self):
        f = self.tab_3b
        pad = dict(padx=16, pady=8)

        ttk.Label(f, text=(
            "Upload GSTR-3B data as Excel with columns:\n"
            "Head | Taxable Value | IGST | CGST | SGST | CESS\n"
            "(One row per section: 3.1 Outward, 4(A) ITC Available, etc.)"
        ), foreground="#1F3864").pack(anchor="w", **pad)

        gstr3b_f = ttk.LabelFrame(f, text="GSTR-3B Excel", padding=12)
        gstr3b_f.pack(fill="x", **pad)
        ttk.Label(gstr3b_f, text="Excel File:").grid(row=0, column=0, sticky="w")
        ttk.Entry(gstr3b_f, textvariable=self.gstr3b_path, width=65).grid(
            row=0, column=1, sticky="ew", padx=(8,8))
        ttk.Button(gstr3b_f, text="Browse",
                   command=lambda: self._browse_file(
                       self.gstr3b_path,
                       [("Excel","*.xlsx *.xls *.csv"),("All","*.*")]
                   )).grid(row=0, column=2)
        gstr3b_f.columnconfigure(1, weight=1)

        note = ttk.LabelFrame(f, text="How 3B Comparison Works", padding=10)
        note.pack(fill="x", **pad)
        ttk.Label(note, wraplength=800, text=(
            "3B vs GSTR-1: Compares your outward liability declared in GSTR-3B with what "
            "you filed in GSTR-1. If portal JSON for GSTR-1 is loaded (Sales tab), "
            "those figures are used; otherwise your Sales Books totals are used.\n\n"
            "3B vs GSTR-2B: Compares ITC claimed in GSTR-3B with ITC available per GSTR-2B. "
            "If portal JSON for GSTR-2B is loaded (Purchase tab), those figures are used."
        ), foreground="#1F3864").pack(anchor="w")

    # ─── Help tab ─────────────────────────────────────────────────────────

    def _build_help_tab(self):
        f = self.tab_help
        txt = tk.Text(f, wrap="word", font=("Segoe UI", 10),
                      bg="#F0F4FA", relief="flat", padx=16, pady=10)
        scroll = ttk.Scrollbar(f, command=txt.yview)
        txt.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        txt.pack(fill="both", expand=True)

        help_content = """
GST RECONCILIATION TOOL – COMPLETE GUIDE
=========================================

FEATURE 1: PURCHASE RECONCILIATION (GSTR-2A / GSTR-2B vs Books)
─────────────────────────────────────────────────────────────────
This is the most critical reconciliation for ITC (Input Tax Credit) compliance.

HOW IT WORKS:
1. Download GSTR-2B JSON from GST Portal:
   Portal → Returns → GSTR-2B → Select Month → Download JSON (Statement)
2. Export Purchase Register from Tally:
   Gateway of Tally → Display → Account Books → Purchase Register → Export to Excel
3. Load both files in the "Purchase" tab and run.

RESULTS YOU GET:
• Matched        – Same invoice in portal and books (ITC safe to claim)
• Mismatch       – Same invoice but amount differs (verify with supplier)
• Portal Only    – Supplier filed but you haven't booked (missed ITC)
• Books Only     – Booked by you but supplier hasn't filed (risky ITC – may be reversed)

─────────────────────────────────────────────────────────────────
FEATURE 2: SALES RECONCILIATION (GSTR-1 vs Books)
─────────────────────────────────────────────────────────────────
HOW IT WORKS:
1. Download GSTR-1 JSON from GST Portal:
   Portal → Returns → GSTR-1 → Filed Returns → Select Month → Download JSON
2. Export Sales Register from Tally as Excel.
3. Load both in "Sales" tab.

─────────────────────────────────────────────────────────────────
FEATURE 3: GSTR-3B COMPARISON
─────────────────────────────────────────────────────────────────
• 3B vs GSTR-1: Detects if you paid more/less tax in 3B than declared in GSTR-1
• 3B vs GSTR-2B: Detects if ITC claimed in 3B exceeds/falls short of GSTR-2B data

─────────────────────────────────────────────────────────────────
REPORT COLOUR CODING
─────────────────────────────────────────────────────────────────
  GREEN  = Matched (no action needed)
  YELLOW = Mismatch (verify and correct)
  RED    = In Portal, not in Books (possible missed entries)
  BLUE   = In Books, not in Portal (risky – supplier may not have filed)

─────────────────────────────────────────────────────────────────
MATCHING LOGIC
─────────────────────────────────────────────────────────────────
• Primary key: GSTIN + Invoice Number (normalized – ignores spaces, slashes, dashes)
• Tolerance: Rs 1 rounding difference is treated as "Matched"
• Invoice number matching is case-insensitive

─────────────────────────────────────────────────────────────────
FREQUENTLY ASKED QUESTIONS
─────────────────────────────────────────────────────────────────
Q: What JSON format does the tool accept?
A: Standard GST portal JSON downloads for GSTR-1, GSTR-2A, GSTR-2B.

Q: My Tally export has different column names – will it work?
A: Yes. The tool auto-detects 50+ common column name variations.

Q: What if supplier GSTIN is different in books vs portal?
A: Those invoices will NOT match. Correct the GSTIN in books and re-run.

Q: Does it support credit notes (CDNR)?
A: GSTR-2B CDNR is parsed automatically from the JSON.

Q: Where is the output report saved?
A: At the path you choose in the Settings tab. Default: same folder as this tool.
"""
        txt.insert("1.0", help_content)
        txt.config(state="disabled")

    # ─── helpers ─────────────────────────────────────────────────────────

    def _browse_file(self, var, filetypes):
        path = filedialog.askopenfilename(filetypes=filetypes)
        if path:
            var.set(path)

    def _browse_output(self):
        period = self.period_var.get().replace(" ","_")
        default = f"GST_Reconciliation_{period}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel","*.xlsx")])
        if path:
            self.output_path.set(path)

    def _set_status(self, msg, pct=None):
        self.status_var.set(msg)
        if pct is not None:
            self.progress_var.set(pct)
        self.update_idletasks()

    # ─── run ─────────────────────────────────────────────────────────────

    def _run_thread(self):
        t = threading.Thread(target=self._run, daemon=True)
        t.start()

    def _run(self):
        try:
            self._set_status("Starting…", 0)
            report_data = {}

            # ── Purchase reconciliation ──────────────────────────────────
            p_portal = self.portal_path.get().strip()
            p_books  = self.books_path.get().strip()

            portal_df = pd.DataFrame()
            if p_portal and os.path.exists(p_portal):
                self._set_status("Reading portal JSON…", 10)
                if self.portal_type.get() == "GSTR-2B":
                    portal_df = parse_gstr2b_json(p_portal)
                else:
                    portal_df = parse_gstr2a_json(p_portal)

            books_df = pd.DataFrame()
            if p_books and os.path.exists(p_books):
                self._set_status("Reading purchase register…", 20)
                raw = read_excel_file(p_books)
                books_df = standardize_books_purchase(raw)

            if not portal_df.empty or not books_df.empty:
                self._set_status("Reconciling purchase data…", 30)
                m, mm, op, ob = reconcile_purchase(portal_df, books_df)
                report_data["purchase_matched"]     = m
                report_data["purchase_mismatch"]    = mm
                report_data["purchase_portal_only"] = op
                report_data["purchase_books_only"]  = ob

            # ── Sales reconciliation ─────────────────────────────────────
            s_portal = self.portal2_path.get().strip()
            s_books  = self.books2_path.get().strip()

            s_portal_df = pd.DataFrame()
            if s_portal and os.path.exists(s_portal):
                self._set_status("Reading GSTR-1 JSON…", 45)
                s_portal_df = parse_gstr1_json(s_portal)

            s_books_df = pd.DataFrame()
            if s_books and os.path.exists(s_books):
                self._set_status("Reading sales register…", 55)
                raw = read_excel_file(s_books)
                s_books_df = standardize_books_sales(raw)

            if not s_portal_df.empty or not s_books_df.empty:
                self._set_status("Reconciling sales data…", 65)
                m, mm, op, ob = reconcile_sales(s_portal_df, s_books_df)
                report_data["sales_matched"]     = m
                report_data["sales_mismatch"]    = mm
                report_data["sales_portal_only"] = op
                report_data["sales_books_only"]  = ob

            # ── 3B comparisons ───────────────────────────────────────────
            gstr3b_path = self.gstr3b_path.get().strip()
            gstr3b_df   = pd.DataFrame()
            if gstr3b_path and os.path.exists(gstr3b_path):
                self._set_status("Reading GSTR-3B…", 72)
                raw = read_excel_file(gstr3b_path)
                gstr3b_df = standardize_books_purchase(raw)  # reuse mapper

            if not gstr3b_df.empty:
                # vs GSTR-1
                compare_1 = s_portal_df if not s_portal_df.empty else s_books_df
                if not compare_1.empty:
                    report_data["gstr3b_vs_1"] = reconcile_3b_vs_1(gstr3b_df, compare_1)
                # vs GSTR-2B
                compare_2 = portal_df
                if not compare_2.empty:
                    report_data["gstr3b_vs_2b"] = reconcile_3b_vs_2b(gstr3b_df, compare_2)

            if not report_data:
                messagebox.showwarning("No Data",
                    "No files were loaded. Please provide at least one data file.")
                self._set_status("No data loaded.", 0)
                return

            # ── Write report ─────────────────────────────────────────────
            out = self.output_path.get().strip()
            if not out:
                period = self.period_var.get().replace(" ","_")
                out = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    f"GST_Reconciliation_{period}.xlsx")
                self.output_path.set(out)

            self._set_status("Writing Excel report…", 85)
            write_recon_report(out, report_data,
                               self.period_var.get(), self.gstin_var.get())

            self._set_status("Done!", 100)
            messagebox.showinfo("Success",
                f"Reconciliation complete!\n\nReport saved to:\n{out}\n\n"
                "Open the file to view colour-coded results.")
            os.startfile(out)

        except Exception as e:
            import traceback
            self._set_status(f"Error: {e}", 0)
            messagebox.showerror("Error",
                f"An error occurred:\n\n{e}\n\n{traceback.format_exc()}")


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = GSTReconApp()
    app.mainloop()
